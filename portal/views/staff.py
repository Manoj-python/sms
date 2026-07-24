"""Internal staff login + audit report. Deliberately separate from the
customer OTP flow in auth.py — no shared cookie, no shared session store,
no self-service signup (accounts provisioned via manage.py create_staff_user).
"""

import io
import re
from datetime import timedelta, timezone as dt_timezone

from django.contrib.auth.hashers import check_password
from django.core.cache import cache
from django.http import HttpResponse, HttpResponseRedirect
from django.shortcuts import render

from portal.decorators import require_staff_session
from portal.models import AuditLog, StaffUser
from portal.services import blocked_loans_service, staff_session_service

IST = dt_timezone(timedelta(hours=5, minutes=30))

# Every audit() call site writes `detail` as space-separated key=value
# tokens (see payments.py/auth.py/dashboard.py) — this is NOT a structured
# field, so it's parsed back out here on read rather than adding a schema
# migration just for report display. Only whichever keys a given action
# actually logged come back; anything else is legitimately absent, not a
# parse failure — see the "Total Dues" comment on _report_row below for why
# that field is never in here.
_DETAIL_FIELD_RE = re.compile(r"(\w+)=(\S+)")

# Superset of every `action` string any audit() call site has ever used
# (see auth.py, dashboard.py, decorators.py, payments.py) — kept as one
# list here (rather than derived from the DB) so the filter dropdown always
# offers every action even before that action has fired once in this DB.
ACTIONS = [
    "login_success", "login_success_agreement", "login_unknown_mobile", "agreement_login_failed",
    "otp_sent", "otp_resent", "logout",
    "qr_generated",
    "statement_downloaded", "statement_viewed",
    "foreclosure_statement_downloaded", "foreclosure_statement_viewed",
    "installment_receipt_downloaded", "installment_receipt_viewed",
    "charge_receipt_downloaded", "charge_receipt_viewed",
    "receipt_downloaded", "receipt_viewed",
    "idor_blocked",
    "download_blocked_dues_contact_requested", "download_blocked_seized_contact_requested",
    "foreclosure_blocked_dues_contact_requested",
]


def _parse_detail(detail: str) -> dict[str, str]:
    return dict(_DETAIL_FIELD_RE.findall(detail or ""))


def _report_row(row: AuditLog) -> dict:
    fields = _parse_detail(row.detail)
    maps_url = None
    if row.latitude is not None and row.longitude is not None:
        maps_url = f"https://www.google.com/maps?q={row.latitude},{row.longitude}"
    return {
        "created_at_ist": row.created_at.astimezone(IST).strftime("%d-%m-%Y %H:%M:%S"),
        "action": row.action,
        "mobile": row.mobile or row.mobile_mask,
        # Loan number: most rows log agreement_no directly; a login row
        # logs every loan the customer holds as "loans=A,B,C" instead
        # (see auth.py) since login isn't scoped to one loan.
        "loan_no": fields.get("agreement_no") or fields.get("loans") or fields.get("agr") or "",
        "emi_due_count": fields.get("emi_due_count", ""),
        # Total Dues is deliberately left blank, not guessed: it was never
        # written into `detail` at the time of the event (only agreement_no
        # /emi_due_count were, and only for a subset of actions), and
        # fetching it live now would show TODAY's dues mislabeled next to a
        # possibly-old historical row. `amount` (present on qr_generated) is
        # the payment amount attempted, not total dues, so it's kept as its
        # own column instead of relabeled. "where ever possible" is
        # satisfied by loan_no/emi_due, which genuinely were captured live.
        "total_dues": "",
        "amount": fields.get("amount", ""),
        "detail": row.detail,
        "ip": row.ip,
        "location": row.location,
        "maps_url": maps_url,
        # Raw coordinates alongside the resolved location text — the HTML
        # report only links out to Google Maps (see maps_url above), but
        # the Excel export needs the actual numbers so staff can paste them
        # into another tool or verify precision without a round-trip
        # through Maps. Blank (not 0) when unresolved, matching maps_url's
        # own None check.
        "latitude": row.latitude if row.latitude is not None else "",
        "longitude": row.longitude if row.longitude is not None else "",
    }


def _filtered_queryset(request):
    action_filter = request.GET.get("action", "")
    mobile_filter = request.GET.get("mobile", "").strip()
    qs = AuditLog.objects.all().order_by("-created_at")
    if action_filter:
        qs = qs.filter(action=action_filter)
    if mobile_filter:
        qs = qs.filter(mobile_mask__icontains=mobile_filter)
    return qs, action_filter, mobile_filter


def _ip(request) -> str:
    return request.META.get("REMOTE_ADDR", "")


def login_page(request):
    expired = request.GET.get("expired")
    return render(request, "staff_login.html", {"expired": expired, "error": None})


async def login_submit(request):
    # Small fixed-window brute-force guard, mirroring ratelimit.py's
    # approach but kept local — staff login has no i18n/error_key contract
    # to share with the customer-facing rate_limit decorator.
    ip = _ip(request)
    key = f"staff_login_attempts:{ip}"
    attempts = cache.get(key, 0)
    if attempts >= 10:
        return render(request, "staff_login.html", {"error": "Too many attempts. Try again later."}, status=429)

    username = (request.POST.get("username") or "").strip()
    password = request.POST.get("password") or ""
    user = await StaffUser.objects.filter(username=username, is_active=True).afirst()
    if user is None or not check_password(password, user.password_hash):
        cache.set(key, attempts + 1, timeout=900)
        return render(request, "staff_login.html", {"error": "Invalid username or password."})

    cache.delete(key)
    sess = await staff_session_service.create_session(username)
    response = HttpResponseRedirect("/staff/report")
    staff_session_service.set_session_cookie(response, sess.id)
    return response


@require_staff_session
async def logout(request, staff):
    await staff_session_service.revoke(staff)
    response = HttpResponseRedirect("/staff/login")
    staff_session_service.clear_session_cookie(response)
    return response


async def login_dispatch(request):
    if request.method == "POST":
        return await login_submit(request)
    return login_page(request)


@require_staff_session
async def report(request, staff):
    """Audit report: logins (OTP + loan-number flow, with the loan numbers
    captured at login time), payment attempts, and receipt/statement
    downloads. See models.AuditLog — `location`/`latitude`/`longitude` are
    filled in asynchronously after each row is written, so they may be
    blank for the most recent minute of activity.

    Mobile numbers are shown unmasked here (an explicit product decision —
    staff need to identify/contact customers from this report) even though
    they're masked everywhere customer-facing. `mobile` is encrypted at
    rest and decrypts automatically via EncryptedCharField; `mobile_mask`
    is the search key since the DB can't filter on encrypted plaintext, and
    is also the display fallback for rows written before `mobile` existed."""
    page = max(1, int(request.GET.get("page", 1) or 1))
    page_size = 50

    qs, action_filter, mobile_filter = _filtered_queryset(request)
    total = await qs.acount()
    rows = [row async for row in qs[(page - 1) * page_size : page * page_size]]
    display_rows = [_report_row(row) for row in rows]

    return render(request, "staff_report.html", {
        "rows": display_rows, "total": total, "page": page, "page_size": page_size,
        "has_next": page * page_size < total, "has_prev": page > 1,
        "actions": ACTIONS, "action_filter": action_filter, "mobile_filter": mobile_filter,
    })


# Hard ceiling on a single export — protects both the request (openpyxl
# building tens of thousands of rows in-process) and Excel itself. Staff
# narrow the filter (action/mobile/date range) rather than get a truncated
# file silently; the export is capped, never trimmed-without-notice.
MAX_EXPORT_ROWS = 20_000


@require_staff_session
async def export(request, staff):
    """Same filters as report(), but the full filtered result set (not just
    the current page) as a downloadable .xlsx — staff want one file to sort
    /pivot in Excel, not 50-row-at-a-time pagination."""
    from openpyxl import Workbook
    from openpyxl.styles import Font
    from openpyxl.utils import get_column_letter

    qs, action_filter, mobile_filter = _filtered_queryset(request)
    total = await qs.acount()
    rows = [row async for row in qs[:MAX_EXPORT_ROWS]]
    display_rows = [_report_row(row) for row in rows]

    wb = Workbook()
    ws = wb.active
    ws.title = "Audit Report"

    headers = ["Date / Time (IST)", "Action", "Mobile", "Loan No.", "EMI Due Count",
               "Amount", "Detail", "IP", "Location", "Latitude", "Longitude"]
    ws.append(headers)
    for cell in ws[1]:
        cell.font = Font(bold=True)

    for r in display_rows:
        ws.append([
            r["created_at_ist"], r["action"], r["mobile"], r["loan_no"],
            r["emi_due_count"], r["amount"], r["detail"], r["ip"], r["location"],
            r["latitude"], r["longitude"],
        ])

    widths = [19, 26, 14, 16, 14, 10, 46, 15, 22, 12, 12]
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.freeze_panes = "A2"

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)

    # Filters land in a Content-Disposition header value below — strip to a
    # safe filename charset rather than trusting the querystring verbatim.
    safe = lambda s: re.sub(r"[^A-Za-z0-9]+", "", s)[:30]
    suffix = ""
    if action_filter:
        suffix += f"_{safe(action_filter)}"
    if mobile_filter:
        suffix += f"_{safe(mobile_filter)}"
    filename = f"audit_report{suffix}.xlsx"

    response = HttpResponse(
        buf.getvalue(),
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )
    response["Content-Disposition"] = f'attachment; filename="{filename}"'
    if total > MAX_EXPORT_ROWS:
        response["X-Export-Truncated"] = f"{MAX_EXPORT_ROWS} of {total}"
    return response


# --- Blocked loans (per-loan access deny-list, see models.BlockedLoan) -----

BLOCKED_LOANS_PAGE_SIZE = 50


@require_staff_session
async def blocked_loans_list(request, staff):
    """View/search the deny-list — every non-empty search field ANDs
    together (see blocked_loans_service.search)."""
    page = max(1, int(request.GET.get("page", 1) or 1))
    qs, filters = blocked_loans_service.search(request)
    total = await qs.acount()
    rows = [
        row async for row in
        qs[(page - 1) * BLOCKED_LOANS_PAGE_SIZE: page * BLOCKED_LOANS_PAGE_SIZE]
    ]
    return render(request, "staff_blocked_loans.html", {
        "rows": rows, "total": total, "page": page,
        "has_next": page * BLOCKED_LOANS_PAGE_SIZE < total, "has_prev": page > 1,
        **filters,
    })


@require_staff_session
async def blocked_loan_add_page(request, staff):
    """Just the agreement-number lookup step — see blocked_loan_lookup for
    the auto-filled form it swaps in."""
    return render(request, "staff_blocked_loan_add.html", {})


@require_staff_session
async def blocked_loan_lookup(request, staff):
    """HTMX: looks up the agreement number live against AllCloud (every
    tenant) and re-renders the SAME add page's form box, either with an
    error or with customer name/mobile/branch pre-filled and editable —
    centre/RSP/reason are always blank (AllCloud has no such fields, see
    models.BlockedLoan)."""
    agreement_no = (request.POST.get("agreement_no") or "").strip().upper()
    found = await blocked_loans_service.lookup_for_add(agreement_no) if agreement_no else None
    if found is None:
        return render(request, "partials/staff_blocked_loan_form.html", {
            "error": "No loan found for that agreement number in any tenant.",
            "agreement_no": agreement_no,
        })
    already = await blocked_loans_service.is_blocked(found["finance_id"], found["lender"])
    return render(request, "partials/staff_blocked_loan_form.html", {"found": found, "already_blocked": already})


@require_staff_session
async def blocked_loan_create(request, staff):
    finance_id = request.POST.get("finance_id", "")
    lender = request.POST.get("lender", "smsquare")
    # The UniqueConstraint on (finance_id, lender) would otherwise raise an
    # unhandled IntegrityError if staff double-submit or re-block an
    # already-blocked loan (the lookup step already warns "already_blocked",
    # but nothing stops them clicking through anyway) — a no-op redirect is
    # a friendlier outcome than a 500 for a harmless double-click.
    if not await blocked_loans_service.is_blocked(finance_id, lender):
        await blocked_loans_service.create(
            finance_id=finance_id,
            lender=lender,
            agreement_no=request.POST.get("agreement_no", ""),
            customer_name=request.POST.get("customer_name", "").strip(),
            mobile=request.POST.get("mobile", "").strip(),
            branch=request.POST.get("branch", "").strip(),
            centre=request.POST.get("centre", "").strip(),
            rsp_name=request.POST.get("rsp_name", "").strip(),
            reason=request.POST.get("reason", "").strip(),
            blocked_by=staff.username,
        )
    return HttpResponseRedirect("/staff/blocked-loans")


@require_staff_session
async def blocked_loan_remove(request, staff):
    await blocked_loans_service.delete(request.POST.get("id", ""))
    return HttpResponseRedirect("/staff/blocked-loans")


@require_staff_session
async def blocked_loans_export(request, staff):
    """Same search filters as the list page, full filtered set as .xlsx —
    the "download for verification" counterpart to bulk upload: block a
    batch, download, and diff against the source sheet."""
    from openpyxl import Workbook
    from openpyxl.styles import Font
    from openpyxl.utils import get_column_letter

    qs, filters = blocked_loans_service.search(request)
    rows = [row async for row in qs[:MAX_EXPORT_ROWS]]

    wb = Workbook()
    ws = wb.active
    ws.title = "Blocked Loans"
    headers = ["Blocked On (IST)", "Agreement No", "Finance Id", "Lender", "Customer Name",
               "Mobile", "Branch", "Centre", "RSP Name", "Reason", "Blocked By"]
    ws.append(headers)
    for cell in ws[1]:
        cell.font = Font(bold=True)
    for r in rows:
        ws.append([
            r.created_at.astimezone(IST).strftime("%d-%m-%Y %H:%M:%S"),
            r.agreement_no, r.finance_id, r.lender, r.customer_name,
            r.mobile or r.mobile_mask, r.branch, r.centre, r.rsp_name, r.reason, r.blocked_by,
        ])
    widths = [19, 22, 12, 12, 24, 14, 16, 16, 16, 40, 14]
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.freeze_panes = "A2"

    buf = io.BytesIO()
    wb.save(buf)
    response = HttpResponse(
        buf.getvalue(),
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )
    response["Content-Disposition"] = 'attachment; filename="blocked_loans.xlsx"'
    return response


# Bulk upload ceiling — every row costs a live cross-tenant AllCloud lookup
# (~1s each even with the concurrency cap below), so a full 500-row sheet
# takes ~2-3 minutes; anything bigger risks blowing request/proxy timeouts
# long before the DB was the constraint (production nginx defaults to a
# 60s proxy_read_timeout — bump that alongside this if uploads 502 there).
# Staff split bigger batches across multiple files.
MAX_UPLOAD_ROWS = 500
# Concurrent AllCloud lookups during an upload — enough to make a full
# 100-row sheet finish in ~20-30s, low enough not to hammer AllCloud (each
# lookup already fans out to all 3 tenants internally).
UPLOAD_LOOKUP_CONCURRENCY = 5


@require_staff_session
async def blocked_loans_sample(request, staff):
    """Downloadable template for bulk upload — generated here (not a static
    file) so its header row can never drift from what _parse_upload
    actually accepts."""
    from openpyxl import Workbook
    from openpyxl.styles import Font
    from openpyxl.utils import get_column_letter

    wb = Workbook()
    ws = wb.active
    ws.title = "Blocked Loans Upload"
    ws.append(["Agreement No", "Centre", "RSP Name", "Reason"])
    for cell in ws[1]:
        cell.font = Font(bold=True)
    # Example rows — obviously-fake agreement numbers, so an accidental
    # upload of the untouched template just reports "not found" per row
    # rather than blocking anything real.
    ws.append(["L2WNXXXXX-000000001", "Example Centre", "Example RSP", "Replace these rows with real data"])
    ws.append(["L2WNXXXXX-000000002", "", "", "Centre / RSP Name / Reason are optional"])
    widths = [24, 18, 18, 42]
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w

    buf = io.BytesIO()
    wb.save(buf)
    response = HttpResponse(
        buf.getvalue(),
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )
    response["Content-Disposition"] = 'attachment; filename="blocked_loans_upload_sample.xlsx"'
    return response


def _parse_upload(uploaded) -> tuple[list[dict], str | None]:
    """Returns (rows, error). Accepts .xlsx (openpyxl) or .csv. The header
    row is required and mapped by name, not position — "agreement"/"loan"
    -> agreement no, "centre"/"center" -> centre, "rsp" -> rsp_name,
    "reason" -> reason — so staff can reuse whatever sheet they already
    have without reordering columns. Unrecognized columns are ignored."""
    import csv

    name = (uploaded.name or "").lower()
    if name.endswith(".csv"):
        try:
            text = uploaded.read().decode("utf-8-sig", errors="replace")
        except Exception:
            return [], "Could not read the CSV file."
        raw_rows = list(csv.reader(text.splitlines()))
    elif name.endswith(".xlsx"):
        from openpyxl import load_workbook
        try:
            wb = load_workbook(uploaded, read_only=True, data_only=True)
        except Exception:
            return [], "Could not read the Excel file — is it a valid .xlsx?"
        ws = wb.active
        raw_rows = [["" if c is None else str(c) for c in row] for row in ws.iter_rows(values_only=True)]
    else:
        return [], "Unsupported file type — upload a .xlsx or .csv file."

    raw_rows = [r for r in raw_rows if any(str(c).strip() for c in r)]
    if not raw_rows:
        return [], "The file is empty."

    header = [str(c).strip().lower() for c in raw_rows[0]]
    col = {}
    for i, h in enumerate(header):
        if ("agreement" in h or "loan" in h) and "agreement_col" not in col:
            col["agreement_col"] = i
        elif "centre" in h or "center" in h:
            col["centre_col"] = i
        elif "rsp" in h:
            col["rsp_col"] = i
        elif "reason" in h:
            col["reason_col"] = i
    if "agreement_col" not in col:
        return [], 'No "Agreement No" (or "Loan No") column found in the header row.'

    def cell(row, key):
        i = col.get(key)
        return str(row[i]).strip() if i is not None and i < len(row) else ""

    rows = []
    for raw in raw_rows[1:]:
        agreement_no = cell(raw, "agreement_col").upper()
        if not agreement_no:
            continue
        rows.append({
            "agreement_no": agreement_no,
            "centre": cell(raw, "centre_col"),
            "rsp_name": cell(raw, "rsp_col"),
            "reason": cell(raw, "reason_col"),
        })
    if not rows:
        return [], "No agreement numbers found under the header row."
    if len(rows) > MAX_UPLOAD_ROWS:
        return [], f"Too many rows ({len(rows)}) — maximum {MAX_UPLOAD_ROWS} per upload. Split the file and retry."
    return rows, None


@require_staff_session
async def blocked_loans_upload(request, staff):
    """GET: upload form + expected-format note. POST: parse the sheet, look
    each agreement number up live against AllCloud (auto-filling customer
    name/mobile/branch exactly like the single-add flow), block what
    resolves, and render a per-row results report — added / already
    blocked / not found — so staff can immediately see what didn't take
    rather than silently dropping rows."""
    import asyncio

    if request.method != "POST":
        return render(request, "staff_blocked_loans_upload.html", {})

    uploaded = request.FILES.get("file")
    if uploaded is None:
        return render(request, "staff_blocked_loans_upload.html", {"error": "Choose a file to upload."})
    rows, error = _parse_upload(uploaded)
    if error:
        return render(request, "staff_blocked_loans_upload.html", {"error": error})

    # Duplicate agreement numbers WITHIN the sheet collapse to one block
    # each — the second occurrence reports "already blocked" naturally via
    # the is_blocked re-check after the first one lands.
    sem = asyncio.Semaphore(UPLOAD_LOOKUP_CONCURRENCY)

    async def resolve(row):
        async with sem:
            found = await blocked_loans_service.lookup_for_add(row["agreement_no"])
        return row, found

    resolved = await asyncio.gather(*(resolve(row) for row in rows))

    results = []
    added = skipped = failed = 0
    for row, found in resolved:
        if found is None:
            results.append({**row, "status": "not_found"})
            failed += 1
            continue
        if await blocked_loans_service.is_blocked(found["finance_id"], found["lender"]):
            results.append({**row, "status": "already_blocked", "customer_name": found["customer_name"]})
            skipped += 1
            continue
        await blocked_loans_service.create(
            finance_id=found["finance_id"], lender=found["lender"],
            agreement_no=found["agreement_no"], customer_name=found["customer_name"],
            mobile=found["mobile"], branch=found["branch"],
            centre=row["centre"], rsp_name=row["rsp_name"], reason=row["reason"],
            blocked_by=staff.username,
        )
        results.append({**row, "status": "added", "customer_name": found["customer_name"]})
        added += 1

    return render(request, "staff_blocked_loans_upload.html", {
        "results": results, "added": added, "skipped": skipped, "failed": failed,
    })
